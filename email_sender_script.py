import smtplib
from email.message import EmailMessage
import openpyxl
import datetime

# Gmail Configuration
EMAIL = "***********"  # Your sender email
PASSWORD = "***************"  # Your generated app password
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587


# Function to send emails
def send_email(recipient, first_name, last_name, gender):
    if not recipient or not last_name or not first_name or not gender:
        print(f"⚠️ Skipping empty or incomplete row: {recipient}, {first_name}, {last_name}, {gender}")
        return False, ""

    # Determine salutation based on gender
    if gender == "F":
        salutation = f"Hi mam {first_name} {last_name}"
    elif gender == "M":
        salutation = f"Hi sir {first_name} {last_name}"
    else:
        salutation = f"Hi {first_name} {last_name}"

    # Email content
    email_content = f"""{salutation},

THE AUTOMATED EMAIL SENDER SEND YOU THIS MAIL FOR TESTING

THANK YOU ☺
"""

    # Create the email message
    msg = EmailMessage()
    msg['From'] = EMAIL
    msg['To'] = recipient
    msg['Subject'] = "Testing"
    msg.set_content(email_content)

    # Send the email
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()  # Upgrade to secure connection
            server.login(EMAIL, PASSWORD)  # Log in to the server
            server.send_message(msg)  # Send the email
        print(f"✅ Email successfully sent to {recipient}")
        return True, email_content  # Return success and message content
    except Exception as e:
        print(f"❌ Failed to send email to {recipient}: {e}")
        return False, ""


# Function to send emails and update Excel
def send_emails_from_excel():
    file_name = "email_sender_datasource.xlsx"  # Ensure this file exists
    
    try:
        wb = openpyxl.load_workbook(file_name)
        customer_sheet = wb["Customer Data"]
        sent_sheet_name = "Sent Emails Data"

        # Ensure "Status" column exists in "Customer Data"
        headers = [cell.value for cell in customer_sheet[1]]
        if "Status" not in headers:
            customer_sheet.cell(row=1, column=len(headers) + 1, value="Status")
            status_col = len(headers) + 1
        else:
            status_col = headers.index("Status") + 1  # Find the column number for "Status"

        # Check if "Sent Emails Data" exists, create it if not
        if sent_sheet_name not in wb.sheetnames:
            sent_sheet = wb.create_sheet(sent_sheet_name)
            sent_sheet.append(["Email", "Sent Date", "Sent Time", "Sent Message"])
        else:
            sent_sheet = wb[sent_sheet_name]

        # Get list of already sent emails to prevent duplicates
        sent_emails = set(sent_sheet.cell(row=i, column=1).value for i in range(2, sent_sheet.max_row + 1))

        sent_count = 0

        for row in customer_sheet.iter_rows(min_row=2, max_col=5):  # Email, Last Name, First Name, Gender, Status
            email_cell, last_name_cell, first_name_cell, gender_cell, status_cell = row

            recipient = email_cell.value
            last_name = last_name_cell.value
            first_name = first_name_cell.value
            gender = gender_cell.value
            status = status_cell.value

            # ✅ SKIP EMPTY ROWS
            if not recipient or not last_name or not first_name or not gender:
                print(f"⚠️ Skipping empty row: {recipient}, {first_name}, {last_name}, {gender}")
                continue

            # ✅ SKIP ALREADY SENT EMAILS
            if status == "Sent":
                print(f"⚠️ Email to {recipient} was already sent. Skipping.")
                continue

            success, message_content = send_email(recipient, first_name, last_name, gender)
            if success:
                sent_date = datetime.date.today().isoformat()
                sent_time = datetime.datetime.now().time().strftime("%H:%M:%S")

                # Update "Customer Data" with "Sent" status
                status_cell.value = "Sent"

                # ✅ Find the first completely empty row in "Sent Emails Data"
                next_row = 2  # Start from the first data row
                while any(sent_sheet.cell(row=next_row, column=col).value for col in range(1, 5)):  # Check all columns
                    next_row += 1  # Move to the next available row

                # ✅ Insert new email data in the first empty row
                sent_sheet.cell(row=next_row, column=1, value=recipient)
                sent_sheet.cell(row=next_row, column=2, value=sent_date)
                sent_sheet.cell(row=next_row, column=3, value=sent_time)
                sent_sheet.cell(row=next_row, column=4, value=message_content)

                sent_count += 1
                if sent_count >= 4:  # Limit to 4 emails for testing
                    break

        # Save the file after formatting
        wb.save(file_name)
        print("✅ Updated 'Customer Data' and logged sent emails in 'Sent Emails Data'.")

    except FileNotFoundError:
        print(f"🚨 Error: The file '{file_name}' was not found.")
    except PermissionError:
        print(f"🚨 Error: The file '{file_name}' is open. Please close it and try again.")

# Run script
if __name__ == "__main__":
    send_emails_from_excel()
